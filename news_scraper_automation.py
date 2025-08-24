#!/usr/bin/env python
# coding: utf-8

# # 📰 Automated BBC Tech News Scraper with Charts & Scheduling

# In[1]:


import pandas as pd
import matplotlib.pyplot as plt
from collections import Counter
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import requests
from bs4 import BeautifulSoup


# In[2]:


def generate_news_report():
    url = "https://www.bbc.com/news/technology"
    response = requests.get(url)


# In[3]:


url = "https://www.bbc.com/news/technology"


# In[4]:


response = requests.get(url)


# In[5]:


soup = BeautifulSoup(response.text, "html.parser")


# In[6]:


headlines = [h.text.strip() for h in soup.find_all("h2")][:20]


# In[7]:


df = pd.DataFrame(headlines, columns=["Headline"] )


# In[8]:


# 2. Keyword analysis


# In[9]:


stopwords = {"the","a","in","of","to","and","on","for","with","by","an","at","as"}
words = []


# In[10]:


for headline in headlines:
    tokens = re.findall(r'\b\w+\b', headline.lower())
    words.extend([w for w in tokens if w not in stopwords and len(w) > 2])


# In[11]:


word_counts = Counter(words).most_common(10)


# In[12]:


df_keywords = pd.DataFrame(word_counts, columns=["Keyword","Count"] )


# In[13]:


# 3. Chart


# In[14]:


plt.figure(figsize=(8,5))
plt.bar(df_keywords["Keyword"], df_keywords["Count"])
plt.title("Top Keywords in BBC Tech Headlines")
plt.xticks(rotation=45)
plt.tight_layout()
chart_file = "keyword_chart.png"
plt.savefig(chart_file)
plt.close()


# In[15]:


# 4. Save Excel


# In[16]:


output_file = f"news_report_{datetime.today().strftime('%Y%m%d')}.xlsx"
df.to_excel(output_file, index=False, sheet_name="Headlines")
with pd.ExcelWriter(output_file, mode="a", engine="openpyxl") as writer:
     df_keywords.to_excel(writer, sheet_name="Keywords", index=False)


# In[17]:


wb = load_workbook(output_file)
ws = wb["Keywords"]
img = Image(chart_file)
ws.add_image(img, "E5")
wb.save(output_file)


# In[18]:


print(f"✅ News report created: {output_file}")


# In[19]:


from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# In[20]:


# Load workbook
wb = load_workbook(output_file)


# In[21]:


# --- Format Headlines sheet ---
ws1 = wb["Headlines"]


# In[22]:


# Bold header
for cell in ws1[1]:
    cell.font = Font(bold=True)


# In[23]:


# Autofit columns
for col in ws1.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws1.column_dimensions[col_letter].width = max_length + 2


# In[24]:


# Alternate row shading
fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
for row in range(2, ws1.max_row + 1, 2):
    for cell in ws1[row]:
        cell.fill = fill


# In[25]:


# --- Format Keywords sheet ---
ws2 = wb["Keywords"]
for cell in ws2[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")


# In[26]:


# Save styled workbook
styled_file = output_file.replace(".xlsx", "_styled.xlsx")
wb.save(styled_file)

print(f"✨ Styled report saved as {styled_file}")


# In[ ]:




